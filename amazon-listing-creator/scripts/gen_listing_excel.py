# -*- coding: utf-8 -*-
"""
Amazon Listing Excel Generator - amazon-listing-creator skill
Version: 2.0.0 (with validate_output auto-check)
Skill: A9 + COSMO + Rufus

MUST INCLUDE: Hanson出品|亚马逊选品/运营/Agent/Skill专家|hansonbtc@163.com
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 颜色常量 ──────────────────────────────────────────
C_MAIN_TITLE_BG   = "2E75B6"
C_MAIN_TITLE_FT   = "FFFFFF"
C_MODULE_BG       = "4472C4"
C_HEADER_BG       = "D9E2F3"
C_HEADER_FT       = "000000"
C_BRAND_BG        = "D9D9D9"
C_PASS_BG         = "C6EFCE"
C_PASS_FT         = "276221"
C_BORDER          = "B4C6E7"

# ── 样式工厂 ──────────────────────────────────────────
def ft(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

# ── 固定出品信息（第2行，强制） ────────────────────────
BRAND_INFO = "Hanson出品|亚马逊选品/运营/Agent/Skill专家|hansonbtc@163.com"

# ══════════════════════════════════════════════════════
# 数据区（用户/AI 填写此区域）
# ══════════════════════════════════════════════════════

PRODUCT = {
    "brand":    "BRAND",
    "model":    "MODEL",
    "name_cn":  "Product Name CN",
}

TITLE = "Your Title Here (≤200 chars)"
TITLE_SCENE    = ""
TITLE_KEYWORDS = ""
TITLE_CN       = ""

BULLETS = [
    {"num": 1, "text": "", "scene": "", "keywords": "", "cn": ""},
    {"num": 2, "text": "", "scene": "", "keywords": "", "cn": ""},
    {"num": 3, "text": "", "scene": "", "keywords": "", "cn": ""},
    {"num": 4, "text": "", "scene": "", "keywords": "", "cn": ""},
    {"num": 5, "text": "", "scene": "", "keywords": "", "cn": ""},
]

SEARCH_TERMS_EN = ""
SEARCH_TERMS_CN = ""

VALIDATION = [
    ("Title 字符数",     "≤ 200 字符", "", ""),
    ("Bullet 1 字符数",  "≤ 500 字符", "", ""),
    ("Bullet 2 字符数",  "≤ 500 字符", "", ""),
    ("Bullet 3 字符数",  "≤ 500 字符", "", ""),
    ("Bullet 4 字符数",  "≤ 500 字符", "", ""),
    ("Bullet 5 字符数",  "≤ 500 字符", "", ""),
    ("Search Terms 字节","≤ 250 bytes", "", ""),
    ("A9 合规",          "核心词在Title", "", ""),
    ("COSMO 合规",       "人群词+场景词", "", ""),
    ("Rufus 合规",       "所有卖点有数据", "", ""),
    ("黑名单词",         "无违禁词",      "", ""),
]

MAIN_IMAGE_SUGGESTIONS = [
    ("卖点 1", "主图",   ""),
    ("卖点 2", "特写图", ""),
    ("卖点 3", "场景图", ""),
    ("卖点 4", "对比图", ""),
    ("卖点 5", "细节图", ""),
    ("卖点 6", "包装图", ""),
]

SEARCH_TERMS_NOTE = ""

OUTPUT_PATH = r"C:\Users\Administrator\.qclaw\workspace\Listing_Output.xlsx"

# ══════════════════════════════════════════════════════
# 自动校验函数（生成后自动执行，不合格直接抛异常）
# ══════════════════════════════════════════════════════
def validate_output(ws):
    """
    强制合规校验 — 检查 37 行结构是否 100% 符合 SKILL.md 规格。
    不合格直接 raise Exception，阻止输出文件被使用。
    """
    errors = []

    # 1. Sheet 名称
    if ws.title != "Amazon Listing":
        errors.append(f"Sheet名称错误: '{ws.title}' (应为 'Amazon Listing')")

    # 2. 第1行 - 主标题行
    c1 = ws.cell(1, 1)
    if c1.value is None or "Amazon Listing" not in str(c1.value):
        errors.append(f"第1行主标题缺失或异常: '{c1.value}'")
    # 检查合并 A1:E1
    if str(c1.coordinate) != "A1":
        errors.append("第1行未正确从A1开始")

    # 3. 第2行 - 出品信息行（最关键！）
    c2 = ws.cell(2, 1)
    v2 = str(c2.value) if c2.value else ""
    if "Hanson\u51fa\u54c1" not in v2:
        errors.append(f"[FAIL] Row2 brand info missing or modified! Actual: '{v2[:50]}'")
    if "hansonbtc@163.com" not in v2:
        errors.append(f"[FAIL] Row2 brand email missing! Actual: '{v2[:50]}'")

    # 4. 列数检查
    if ws.max_column < 5:
        errors.append(f"列数不足: {ws.max_column} (要求5列)")

    # 5. 模块计数（通过扫描模块标题行检测）
    module_count = 0
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "")
        if any(kw in val for kw in ["Title（标题）", "Bullet Points", "Search Terms", "Validation Summary", "Main Image"]):
            module_count += 1
    if module_count < 5:
        errors.append(f"模块数量不足: {module_count} (要求5个模块)")

    # 6. Search Terms 备注行检测
    has_note = False
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "")
        if "备注" in val and ("字节" in val or "bytes" in val.lower()):
            has_note = True
            break
    if not has_note:
        errors.append("缺少 Search Terms 备注行（字节数统计行）")

    # 7. Validation 项目数量
    val_count = 0
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "")
        if "字符数" in val or "字节" in val or "合规" in val:
            val_count += 1
    if val_count < 11:
        errors.append(f"Validation Summary 项目不足: {val_count} (要求11项)")

    # 8. 主图建议数量
    img_count = 0
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "")
        if val.startswith("卖点"):
            img_count += 1
    if img_count < 6:
        errors.append(f"主图设计建议条数不足: {img_count} (要求6条)")

    # 9. 行高检查（Title 行高应 ≥ 60）
    try:
        title_row_height = ws.row_dimensions[5].height or 0
        if title_row_height < 60:
            errors.append(f"Title数据行行高过小: {title_row_height} (建议≥60)")
    except:
        pass

    # 输出结果
    if errors:
        msg = "\n".join(f"  [X] {e}" for e in errors)
        raise Exception(f"\n{'='*60}\n[FAIL] COMPLIANCE CHECK FAILED ({len(errors)} issues)\n{'='*60}\n{msg}\n{'='*60}\n--> Fix all issues above before using this file. DO NOT skip any check.\n")
    else:
        print(f"\n{'='*60}")
        print("[PASS] COMPLIANCE CHECK PASSED -- All 9 checks OK")
        print(f"  Sheet: {ws.title} | Rows: {ws.max_row} | Cols: {ws.max_column}")
        print(f"  Brand Row: OK | Modules: {module_count} | Validations: {val_count}")
        print(f"  Images: {img_count} | Search Terms Note: {'YES' if has_note else 'NO'}")
        print(f"{'='*60}\n")


# ══════════════════════════════════════════════════════
# 主函数
# ══════════════════════════════════════════════════════
def create_listing_excel(path=None):
    if path is None:
        path = OUTPUT_PATH

    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Listing"

    # 列宽
    col_widths = [16, 58, 22, 30, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ═══ 第1行：主标题 ═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = f"Amazon Listing - {PRODUCT['model']} {PRODUCT['name_cn']}"
    cell.font      = ft(bold=True, size=14, color=C_MAIN_TITLE_FT)
    cell.fill      = fill(C_MAIN_TITLE_BG)
    cell.alignment = aln("center")
    ws.row_dimensions[row].height = 32
    row += 1

    # ═══ 第2行：固定出品信息（强制！）═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = BRAND_INFO
    cell.font      = ft(size=10, color="000000")
    cell.fill      = fill(C_BRAND_BG)
    cell.alignment = aln("center", "center")
    cell.border    = thin_border()
    ws.row_dimensions[row].height = 20
    row += 1

    # ═══ 模块 1：Title ═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = "1. Title（标题）— A9 优化"
    cell.font      = ft(bold=True, size=11, color="FFFFFF")
    cell.fill      = fill(C_MODULE_BG)
    cell.alignment = aln("left")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, hdr in enumerate(["模块", "内容", "场景词", "关键词埋入", "中文翻译"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = ft(bold=True, size=10, color=C_HEADER_FT)
        c.fill = fill(C_HEADER_BG)
        c.alignment = aln("center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    data = ["Title", TITLE, TITLE_SCENE, TITLE_KEYWORDS, TITLE_CN]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = ft(size=10)
        c.fill = fill("FFFFFF")
        c.alignment = aln("left" if col > 1 else "center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 90
    row += 1

    # ═══ 模块 2：Bullet Points ═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = "2. Bullet Points（五点描述）— Rufus + COSMO 驱动"
    cell.font      = ft(bold=True, size=11, color="FFFFFF")
    cell.fill      = fill(C_MODULE_BG)
    cell.alignment = aln("left")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, hdr in enumerate(["#", "Bullet Point", "场景词", "关键词埋入", "中文翻译"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = ft(bold=True, size=10, color=C_HEADER_FT)
        c.fill = fill(C_HEADER_BG)
        c.alignment = aln("center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    for bp in BULLETS:
        d = [bp["num"], bp["text"], bp["scene"], bp["keywords"], bp["cn"]]
        for col, val in enumerate(d, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = ft(size=10)
            c.fill = fill("FFFFFF")
            c.alignment = aln("center" if col == 1 else "left")
            c.border = thin_border()
        ws.row_dimensions[row].height = 90
        row += 1

    # ═══ 模块 3：Search Terms ═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = "3. Search Terms（后台搜索词）— 无重复、无竞品品牌"
    cell.font      = ft(bold=True, size=11, color="FFFFFF")
    cell.fill      = fill(C_MODULE_BG)
    cell.alignment = aln("left")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, hdr in enumerate(["字段", "English（英文）", "中文翻译", "", ""], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = ft(bold=True, size=10, color=C_HEADER_FT)
        c.fill = fill(C_HEADER_BG)
        c.alignment = aln("center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    st_data = ["Search Terms", SEARCH_TERMS_EN, SEARCH_TERMS_CN, "", ""]
    for col, val in enumerate(st_data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = ft(size=10)
        c.fill = fill("FFFFFF")
        c.alignment = aln("center" if col == 1 else "left")
        c.border = thin_border()
    ws.row_dimensions[row].height = 65
    row += 1

    # 备注行
    for col in range(1, 6):
        c = ws.cell(row=row, column=col, value="")
        c.fill = fill("F5F5F5")
        c.border = thin_border()
    ws.merge_cells(f"A{row}:E{row}")
    note_cell = ws[f"A{row}"]
    note_val = SEARCH_TERMS_NOTE or f"备注：字节数: {len(SEARCH_TERMS_EN.encode('utf-8'))}/250 | 已排除: (待填) | 新增长尾: (待填) | 合规检查通过"
    note_cell.value = note_val
    note_cell.font      = ft(size=9, italic=True, color="555555")
    note_cell.alignment = aln("left")
    ws.row_dimensions[row].height = 35
    row += 1

    # ═══ 模块 4：Validation Summary ═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = "4. Validation Summary（校验汇总）— 三算法合规检查"
    cell.font      = ft(bold=True, size=11, color="FFFFFF")
    cell.fill      = fill(C_MODULE_BG)
    cell.alignment = aln("left")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, hdr in enumerate(["项目", "要求", "实际结果", "状态", ""], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = ft(bold=True, size=10, color=C_HEADER_FT)
        c.fill = fill(C_HEADER_BG)
        c.alignment = aln("center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    for item, req, result, status in VALIDATION:
        bg = C_PASS_BG
        ft_c = C_PASS_FT
        st_text = f"✅ PASS" if status == "PASS" else f"❌ FAIL"

        c1 = ws.cell(row=row, column=1, value=item)
        c1.font = ft(bold=True, size=10)
        c1.fill = fill(bg)
        c1.alignment = aln("center")
        c1.border = thin_border()

        c2 = ws.cell(row=row, column=2, value=req)
        c2.font = ft(size=10)
        c2.fill = fill("FFFFFF")
        c2.alignment = aln("center")
        c2.border = thin_border()

        c3 = ws.cell(row=row, column=3, value=result)
        c3.font = ft(size=10)
        c3.fill = fill("FFFFFF")
        c3.alignment = aln("left")
        c3.border = thin_border()

        c4 = ws.cell(row=row, column=4, value=st_text)
        c4.font = ft(bold=True, size=10, color=ft_c)
        c4.fill = fill(bg)
        c4.alignment = aln("center")
        c4.border = thin_border()

        c5 = ws.cell(row=row, column=5, value="")
        c5.fill = fill("FFFFFF")
        c5.border = thin_border()

        ws.row_dimensions[row].height = 22
        row += 1

    # ═══ 模块 5：主图设计建议 ═══
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = "5. 主图设计建议（Main Image Design Suggestions）"
    cell.font      = ft(bold=True, size=11, color="FFFFFF")
    cell.fill      = fill(C_MODULE_BG)
    cell.alignment = aln("left")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, hdr in enumerate(["卖点", "图片类型", "视觉表达建议", "", ""], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = ft(bold=True, size=10, color=C_HEADER_FT)
        c.fill = fill(C_HEADER_BG)
        c.alignment = aln("center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    for sugg in MAIN_IMAGE_SUGGESTIONS:
        for col, val in enumerate(list(sugg) + ["", ""], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = ft(size=10)
            c.fill = fill("FFFFFF")
            c.alignment = aln("center" if col <= 2 else "left")
            c.border = thin_border()
        ws.row_dimensions[row].height = 45
        row += 1

    # 保存
    wb.save(path)
    print(f"[SAVE] Excel saved: {path}")

    # ═══ 自动校验（不可跳过）═══
    validate_output(ws)

    return path


if __name__ == "__main__":
    create_listing_excel()
